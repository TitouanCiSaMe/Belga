"""
Export des fiches vers un fichier Excel formaté (.xlsx).
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_HEADERS = [
    "Nom", "Prénom", "N° Carte", "Spécialisation",
    "Adresse", "Code Postal", "Ville", "Journal / Statut",
    "Tél. Bureau", "Tél. Privé", "Fax", "GSM", "Email",
    "Annuaire", "Page",
]
_KEYS = [
    'nom', 'prenom', 'carte_presse', 'specialisation',
    'adresse', 'code_postal', 'ville', 'journal_ou_statut',
    'tel_bureau', 'tel_prive', 'fax', 'gsm', 'email',
    'annuaire', 'page',
]
_WIDTHS = [22, 18, 10, 12, 35, 10, 25, 30, 18, 18, 18, 18, 30, 12, 6]


def fiches_to_excel(fiches: list[dict], output_path) -> None:
    """Exporte une liste de fiches dans un fichier Excel formaté.

    Fonctionnalités : en-tête bleue, lignes alternées, filtres automatiques,
    première ligne figée, largeurs de colonnes ajustées.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Journalistes"

    hf  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    hfi = PatternFill("solid", fgColor="2F5496")
    ha  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cf  = Font(name="Arial", size=10)
    ca  = Alignment(vertical="top", wrap_text=True)
    bd  = Border(*(Side('thin'),) * 4)
    zebra = PatternFill("solid", fgColor="D6E4F0")

    for c, h in enumerate(_HEADERS, 1):
        cell = ws.cell(1, c, h)
        cell.font, cell.fill, cell.alignment, cell.border = hf, hfi, ha, bd

    for ri, f in enumerate(fiches, 2):
        for ci, k in enumerate(_KEYS, 1):
            cell = ws.cell(ri, ci, f.get(k, ''))
            cell.font, cell.alignment, cell.border = cf, ca, bd
            if ri % 2 == 1:
                cell.fill = zebra

    for i, w in enumerate(_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.auto_filter.ref = f"A1:{get_column_letter(len(_HEADERS))}{len(fiches) + 1}"
    ws.freeze_panes = "A2"

    # Masquer la colonne Spécialisation si elle est vide (formats A/B)
    if not any(f.get('specialisation', '') for f in fiches):
        ws.column_dimensions['D'].hidden = True

    wb.save(output_path)
    print(f"Excel exporté : {output_path} ({len(fiches)} fiches)")
