#!/usr/bin/env python3
"""
Parser unifié — Annuaires de la Presse Belge (1954-2024) — v2

Corrections par rapport à v1 :
  1. Refonte de _is_name_b : heuristique renforcée, blacklist de patterns
  2. Lookahead pour noms multi-lignes (format B)
  3. Filtre des renvois "Voir..."
  4. Nettoyage étendu des artefacts OCR (icônes, caractères parasites)
  5. smart_gutter activé aussi pour le format A
  + Rapport qualité, --dry-run, logging, try/except par page

Usage:
    python parser_annuaires_v2.py annuaire_2003.pdf -o fiches.xlsx
    python parser_annuaires_v2.py annuaire_2000.pdf --format b --pages 155-390 -o fiches.xlsx
"""

import pdfplumber, re, argparse, sys, logging
from pathlib import Path
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

log = logging.getLogger("parser_annuaires")


# ═════════════════════════════════════════════════════════════
# 1. CONSTANTES COMMUNES
# ═════════════════════════════════════════════════════════════

PARTICLES = {'de', 'den', 'der', 'van', 'ten', 'ter', 'dit', 'dite',
             'el', 'da', 'di', 'du', 'le', 'la'}

CARTE_RE_AB = re.compile(r'^[FNA][S/]?\d[\dS ]{2,5}$')
CARTE_RE_C  = re.compile(r'[FN]S?\d{4,5}')
CARTE_INLINE_RE = re.compile(r'[FN]S?\d{3,5}')

CP_VILLE_RE = re.compile(
    r'^(\d{4,5})\s+([A-ZÉÈÊËÀÂÙÛÜÔÏÎ][A-ZÉÈÊËÀÂÙÛÜÔÏÎ\s\-\'/\(\)\.]*)'
)

STATUTS = [
    'Pensionné actif', 'Pensionnéactif', 'Honoraire', 'Erelid',
    'Indépendant', 'Zelfstandige', 'Freelance',
    'Actief gepensioneerde', 'Aktief gepensioneerde',
]

SKIP_HEADERS = {'Redacteurs', 'Rédacteurs', 'Journalistes', 'Listes'}

# ── FIX 3 : Pattern renvoi "Voir..." ──
VOIR_RE = re.compile(r'^Voir\s', re.I)

# ── FIX 4 : Artefacts OCR étendus ──
OCR_ICON_TRAIL_RE = re.compile(
    r'[\s]*(?:Ük|[/\\^<>*©®•✎🌐☎📞📠📧🖊🖋✒♂♀§†‡¶€£¥Üü])+\s*$'
)
OCR_SUFFIX_RE = re.compile(
    r'\s+(?:fe|Ük|<\*©•|[/\\^][\?]?)\s*$'
)
OCR_PREFIX_RE = re.compile(
    r'^(?:[/\\^✎🌐<>©®•][\s]*|fe\s+|Ük\s+|j\s+(?=[A-Z]))'
)

# Format C : tags de spécialisation
SPEC_TAGS = {
    'ECO', 'POL', 'INT', 'REG', 'CULT', 'MED', 'SOC', 'SPO',
    'VFD', 'SCI', 'ENV', 'JUS', 'DEF',
}

# Format B : contact patterns
TEL_BU_RE = re.compile(r'^Tél\.\s*bu\.\s*:?\s*(.*)')
TEL_PR_RE = re.compile(r'^Tél\.\s*pr\.\s*:?\s*(.*)')
FAX_RE    = re.compile(r'^Fax\s*:?\s*(.*)')
GSM_RE    = re.compile(r'^GSM\s*:?\s*(.*)')
EMAIL_PAT = re.compile(r'^Email\s*:?\s*(.*)', re.I)
CONTACT_RES = [
    ('tel_bureau', TEL_BU_RE), ('tel_prive', TEL_PR_RE),
    ('fax', FAX_RE), ('gsm', GSM_RE), ('email', EMAIL_PAT),
]

EMAIL_RE = re.compile(r'[\w\.\-]+@[\w\.\-]+')
PHONE_RE = re.compile(r'[\d/\.\-]{7,}')

# ── FIX 1 : Blacklist pour _is_name_b ──
NAME_BLACKLIST_STARTS = re.compile(
    r'^(?:Rue|Avenue|Av\.|Boulevard|Bd|Blvd|Chaussée|Chemin|Drève|Place|Impasse|'
    r'Route|Allée|Square|Quai|Clos|Sentier|Voie|Passage|'
    r'Straat|Laan|Weg|Plein|Steenweg|Dreef|Kaai|Lei|'
    r'Tél|Fax|GSM|Email|Tel\.|'
    r'Voir|See|Zie|'
    r'Pensionn|Honoraire|Erelid|Indépendant|Freelance|Zelfstandige|'
    r'Aktief|Actief|'
    r'l/l|CD|\+->|</l|ru\s|c$|i/l|S—|=3)\b',
    re.I
)


def new_fiche(**kw):
    """Crée une fiche vide avec tous les champs."""
    f = {
        'nom': '', 'prenom': '', 'carte_presse': '', 'specialisation': '',
        'adresse': '', 'code_postal': '', 'ville': '',
        'journal_ou_statut': '',
        'tel_bureau': '', 'tel_prive': '', 'fax': '', 'gsm': '', 'email': '',
    }
    f.update(kw)
    return f


# ═════════════════════════════════════════════════════════════
# 2. UTILITAIRES TEXTE
# ═════════════════════════════════════════════════════════════

def clean_ocr_icons(text):
    """Supprime les artefacts OCR d'icônes en début/fin de ligne."""
    text = OCR_SUFFIX_RE.sub('', text)
    text = OCR_ICON_TRAIL_RE.sub('', text)
    text = OCR_PREFIX_RE.sub('', text)
    return text.strip()


def fix_spacing(text):
    """Corrige les espaces manquants : MERCKXAnnick → MERCKX Annick."""
    text = re.sub(r'([A-ZÉÈÊËÀÂÙÛÜÔÏÎ])([A-Z][a-zéèêëàâùûüôïî])', r'\1 \2', text)
    text = re.sub(r'([a-zéèêëàâùûüôïî,])([A-ZÉÈÊËÀÂÙÛÜÔÏÎ])', r'\1 \2', text)
    text = re.sub(r'(\d)([A-ZÉÈÊËÀÂ])', r'\1 \2', text)
    text = re.sub(r'([a-zéèêëàâùûüôïî])(\d)', r'\1 \2', text)
    return text


def split_nom_prenom(name_str):
    """Sépare 'VAN DEN EYNDE Wim' en ('VAN DEN EYNDE', 'Wim')."""
    words = name_str.split()
    nom, prenom, found = [], [], False
    for w in words:
        stripped = re.sub(r'[^A-Za-zÉÈÊËÀÂÙÛÜÔÏÎéèêëàâùûüôïî\-\'\(\)]', '', w)
        if not stripped:
            continue
        if not found:
            w_core = stripped.replace("'", "").replace("-", "")
            is_upper = w_core.isupper() and len(w_core) >= 1
            is_particle = stripped.lower() in PARTICLES
            if is_upper or is_particle:
                nom.append(w)
            else:
                found = True
                prenom.append(w)
        else:
            prenom.append(w)
    return ' '.join(nom), ' '.join(prenom)


def clean_carte(raw):
    """Nettoie un numéro de carte : 'F438 S' → 'F4385', 'F28 S3' → 'F2853'."""
    # S précédé d'espace et suivi d'un chiffre → OCR de 5+chiffre? Non, garder.
    # S en fin de chaîne précédé d'espace → OCR de 5
    cleaned = re.sub(r'\s+S$', '5', raw)
    # S au milieu entouré d'espaces : 'F28 S3' → 'F2853'
    cleaned = re.sub(r'\s+S(\d)', r'5\1', cleaned)
    return cleaned.replace(' ', '')


def clean_prenom(prenom):
    """Nettoie un prénom des artefacts OCR résiduels."""
    prenom = re.sub(r'\s+(?:\d|Ük|<\*©•|fe|[/\\^<>©®•✎🌐]+[\?]?)$', '', prenom)
    return prenom.strip()


# ═════════════════════════════════════════════════════════════
# 3. EXTRACTION TEXTE (2 colonnes / 1 colonne)
# ═════════════════════════════════════════════════════════════

def words_to_lines(words, y_tol=5):
    """Regroupe les mots par ligne (proximité verticale)."""
    if not words:
        return []
    lines, cur = [], [words[0]]
    for w in words[1:]:
        if abs(w['top'] - cur[-1]['top']) <= y_tol:
            cur.append(w)
        else:
            cur.sort(key=lambda x: x['x0'])
            lines.append(' '.join(x['text'] for x in cur))
            cur = [w]
    cur.sort(key=lambda x: x['x0'])
    lines.append(' '.join(x['text'] for x in cur))
    return lines


def find_gutter(words, page_width):
    """Trouve le x de séparation entre 2 colonnes (milieu du gutter)."""
    real = [w for w in words if len(w['text'].strip()) > 2]
    if not real:
        return page_width / 2
    step = 5
    buckets = Counter()
    for w in real:
        buckets[int(w['x0'] / step) * step] += 1
    min_b = (int(page_width * 0.25) // step) * step
    max_b = (int(page_width * 0.75) // step) * step
    best_start, best_end, best_size = None, None, 0
    gap_start = None
    for b in range(min_b, max_b + step, step):
        if buckets.get(b, 0) == 0:
            if gap_start is None:
                gap_start = b
        else:
            if gap_start is not None:
                size = b - gap_start
                if size > best_size:
                    best_size, best_start, best_end = size, gap_start, b
                gap_start = None
    if best_start is not None and best_size > 15:
        return (best_start + best_end) / 2
    return page_width / 2


def extract_two_columns(page):
    """Extrait le texte d'une page à 2 colonnes.
    
    FIX 5 : smart_gutter activé par défaut pour tous les formats.
    """
    words = page.extract_words(keep_blank_chars=True, x_tolerance=3, y_tolerance=3)
    if not words:
        return []
    mid_x = find_gutter(words, page.width)
    left  = sorted([w for w in words if w['x0'] < mid_x],  key=lambda w: (w['top'], w['x0']))
    right = sorted([w for w in words if w['x0'] >= mid_x], key=lambda w: (w['top'], w['x0']))
    return words_to_lines(left) + words_to_lines(right)


def extract_single_column(page):
    """Extrait le texte d'une page à 1 colonne (format C)."""
    text = page.extract_text(x_tolerance=3, y_tolerance=3)
    if not text:
        return []
    return [l.strip() for l in text.split('\n') if l.strip()]


# ═════════════════════════════════════════════════════════════
# 4. FORMAT A — 1995-1999
#    2 colonnes, carte sur la ligne du nom, contacts simples
# ═════════════════════════════════════════════════════════════

def _detect_name_a(line):
    """Retourne (nom, prenom, carte) ou None."""
    m = CARTE_INLINE_RE.search(line)
    if not m:
        return None
    carte = m.group()
    name_part = line[:m.start()].strip()
    if not name_part or len(name_part) < 3:
        return None
    name_part = clean_ocr_icons(name_part)
    words = name_part.split()
    if len(words) < 2:
        return None
    first = re.sub(r'[^A-Za-zÉÈÊËÀÂÙÛÜÔÏÎéèêëàâùûüôïî\-\'\(\)]', '', words[0])
    if not first or not first[0].isupper():
        return None
    if NAME_BLACKLIST_STARTS.match(name_part):
        return None
    nom, prenom = split_nom_prenom(name_part)
    return nom, prenom, carte


def parse_format_a(lines, debug=False):
    fiches, current, phase = [], None, 'waiting'

    for raw_line in lines:
        line = clean_ocr_icons(fix_spacing(raw_line.strip()))
        if not line or line in SKIP_HEADERS or re.match(r'^\d{1,3}$', line):
            continue
        if VOIR_RE.match(line):
            if debug: print(f"  [SKIP renvoi] {line}")
            continue
        if debug:
            print(f"  [{phase:12s}] {line}")

        parsed = _detect_name_a(line)
        if parsed:
            if current and current['nom']:
                fiches.append(current)
            nom, prenom, carte = parsed
            current = new_fiche(nom=nom, prenom=clean_prenom(prenom), carte_presse=carte)
            phase = 'got_name'
            continue

        if not current:
            continue

        m_cp = CP_VILLE_RE.match(line)
        if m_cp and phase in ('got_name', 'got_addr'):
            current['code_postal'] = m_cp.group(1)
            ville = m_cp.group(2).strip()
            cp2 = re.search(r'\s+\d{4}\s+', ville)
            if cp2:
                ville = ville[:cp2.start()].strip()
            current['ville'] = ville
            phase = 'got_cp'
            continue

        statut_matched = False
        for statut in STATUTS:
            if line.startswith(statut):
                current['journal_ou_statut'] = statut
                phones = PHONE_RE.findall(line[len(statut):])
                if len(phones) >= 1: current['tel_bureau'] = phones[0]
                if len(phones) >= 2: current['tel_prive'] = phones[1]
                phase = 'got_journal'
                statut_matched = True
                break
        if statut_matched:
            continue

        em = EMAIL_RE.search(line)
        if em and phase in ('got_journal', 'got_cp'):
            current['email'] = em.group()
            continue

        if phase == 'got_cp':
            cleaned = CARTE_INLINE_RE.sub('', line).strip()
            phones = PHONE_RE.findall(cleaned)
            journal = cleaned
            for p in phones:
                journal = journal.replace(p, '').strip()
            journal = clean_ocr_icons(journal).strip(' ,')
            current['journal_ou_statut'] = journal
            if len(phones) >= 1: current['tel_bureau'] = phones[0]
            if len(phones) >= 2: current['tel_prive'] = phones[1]
            phase = 'got_journal'
            continue

        if phase in ('got_name', 'got_addr'):
            if not current['adresse']:
                current['adresse'] = line.rstrip(',')
            else:
                current['adresse'] += ', ' + line.rstrip(',')
            phase = 'got_addr'

    if current and current['nom']:
        fiches.append(current)
    return fiches


# ═════════════════════════════════════════════════════════════
# 5. FORMAT B — 2000-2002
#    2 colonnes, carte séparée, contacts labellisés (Tél. bu.)
# ═════════════════════════════════════════════════════════════

def _is_name_b(line):
    """Détecte une ligne-nom format B (NOM Prénom).
    
    FIX 1 : Heuristique renforcée.
    """
    cleaned = clean_ocr_icons(line)
    if not cleaned or len(cleaned) < 4:
        return False
    if NAME_BLACKLIST_STARTS.match(cleaned):
        return False
    if CARTE_RE_AB.match(cleaned) or CP_VILLE_RE.match(cleaned):
        return False
    for _, pat in CONTACT_RES:
        if pat.match(cleaned):
            return False
    if re.match(r'^\d{1,3}$', cleaned) or cleaned in SKIP_HEADERS:
        return False

    words = cleaned.split()
    if len(words) < 2:
        return False

    first_alpha = re.sub(r'[^A-Za-zÉÈÊËÀÂÙÛÜÔÏÎéèêëàâùûüôïî]', '', words[0])
    if not first_alpha or len(first_alpha) < 2 or not first_alpha[0].isupper():
        return False

    # Trouver la frontière nom/prénom
    nom_words = []
    prenom_start = -1
    for i, w in enumerate(words):
        w_alpha = re.sub(r'[^A-Za-zÉÈÊËÀÂÙÛÜÔÏÎéèêëàâùûüôïî\-\']', '', w)
        if not w_alpha:
            continue
        w_core = w_alpha.replace("'", "").replace("-", "")
        if w_core.isupper() or w_alpha.lower() in PARTICLES:
            nom_words.append(w)
        else:
            prenom_start = i
            break

    if not nom_words or prenom_start < 0:
        return False

    prenom_first = re.sub(r'[^A-Za-zÉÈÊËÀÂÙÛÜÔÏÎéèêëàâùûüôïî]', '', words[prenom_start])
    if not prenom_first or not prenom_first[0].isupper():
        return False

    # Le nom doit contenir au moins un mot uppercase NON-particule
    has_real_nom = False
    for nw in nom_words:
        nw_core = re.sub(r'[^A-Za-z]', '', nw)
        if len(nw_core) >= 2 and nw_core.isupper() and nw.lower() not in PARTICLES:
            has_real_nom = True
            break
    if not has_real_nom:
        return False

    # Rejeter si le dernier mot du nom se termine par un tiret (mot coupé par OCR)
    last_nom = nom_words[-1]
    if last_nom.endswith('-'):
        return False

    return True


def parse_format_b(lines, debug=False):
    fiches, current, phase = [], None, 'waiting'

    # Pré-traitement : nettoyage + indexation pour lookahead
    line_list = []
    for raw_line in lines:
        line = fix_spacing(raw_line.strip())
        line = clean_ocr_icons(line)
        if line:
            line_list.append(line)

    i = 0
    while i < len(line_list):
        line = line_list[i]

        if not line or line in SKIP_HEADERS or re.match(r'^\d{1,3}$', line) \
           or re.match(r'^[FNA]-\d+$', line):
            i += 1; continue

        if VOIR_RE.match(line):
            if debug: print(f"  [SKIP renvoi] {line}")
            i += 1; continue

        if debug:
            print(f"  [{phase:12s}] {line}")

        # Contacts labellisés
        contact_matched = False
        for field, pat in CONTACT_RES:
            m = pat.match(line)
            if m and current:
                val = m.group(1).strip()
                if val:
                    current[field] = val
                phase = 'contacts'
                contact_matched = True
                break
        if contact_matched:
            i += 1; continue

        # Email isolé
        if phase == 'contacts' and current and '@' in line and not current.get('email'):
            current['email'] = line.strip()
            i += 1; continue

        # Journal (après CP) — vérifié AVANT la détection de nom
        # pour éviter que "RTL-TVi" ou "La Meuse Luxembourg" soient pris pour des noms
        if phase == 'got_cp' and current:
            current['journal_ou_statut'] = clean_ocr_icons(line)
            phase = 'got_journal'
            i += 1; continue

        # Ligne-nom
        if _is_name_b(line):
            if current and current.get('nom'):
                fiches.append(current)
            name_clean = clean_ocr_icons(line)
            nom, prenom = split_nom_prenom(name_clean)
            prenom = clean_prenom(prenom)
            current = new_fiche(nom=nom, prenom=prenom)
            phase = 'got_name'

            # FIX 2 : Lookahead pour noms multi-lignes
            if i + 1 < len(line_list):
                next_line = line_list[i + 1]
                next_clean = clean_ocr_icons(next_line)
                # La ligne suivante est un complément de prénom si :
                # - elle ne match aucun pattern structurel
                # - elle est courte, commence par une majuscule
                # - la ligne d'après est une carte
                is_structural = (
                    CARTE_RE_AB.match(next_clean)
                    or CP_VILLE_RE.match(next_clean)
                    or any(p.match(next_clean) for _, p in CONTACT_RES)
                    or _is_name_b(next_clean)
                    or VOIR_RE.match(next_clean)
                )
                if (not is_structural
                    and next_clean
                    and next_clean[0].isupper()
                    and not next_clean[0].isdigit()
                    and len(next_clean.split()) <= 3
                    and not re.search(r'\d{4}', next_clean)
                    and not NAME_BLACKLIST_STARTS.match(next_clean)
                    and i + 2 < len(line_list)
                    and CARTE_RE_AB.match(clean_ocr_icons(line_list[i + 2]))):
                    # Complément de prénom confirmé
                    if prenom:
                        current['prenom'] = prenom + ' ' + next_clean
                    else:
                        current['prenom'] = clean_prenom(next_clean)
                    if debug:
                        print(f"  [MULTILINE  ] +prénom: {next_clean}")
                    i += 1

            i += 1; continue

        # FIX 2b : Nom full-uppercase sans prénom sur la même ligne
        # Ex : "ARGUELLES GONZALEZ\nMarcelino\nF3374"
        line_clean = clean_ocr_icons(line)
        if (line_clean
            and not NAME_BLACKLIST_STARTS.match(line_clean)
            and not CARTE_RE_AB.match(line_clean)
            and not CP_VILLE_RE.match(line_clean)
            and len(line_clean) >= 3
            and line_clean.replace(' ', '').replace('-', '').replace("'", '').isalpha()
            and line_clean.upper() == line_clean  # tout uppercase
            and i + 2 < len(line_list)):
            next1 = clean_ocr_icons(line_list[i + 1])
            next2 = clean_ocr_icons(line_list[i + 2])
            # Vérifier: next1 = prénom mixed-case, next2 = carte
            if (next1 and next1[0].isupper() and not next1.isupper()
                and len(next1.split()) <= 3
                and not NAME_BLACKLIST_STARTS.match(next1)
                and CARTE_RE_AB.match(next2)):
                if current and current.get('nom'):
                    fiches.append(current)
                nom = line_clean
                prenom = clean_prenom(next1)
                current = new_fiche(nom=nom, prenom=prenom)
                phase = 'got_name'
                if debug:
                    print(f"  [FULLCAPS   ] {nom} + {prenom}")
                i += 2  # sauter la ligne prénom (la carte sera traitée au tour suivant)
                continue

        if not current:
            i += 1; continue

        # Carte (ligne isolée)
        if CARTE_RE_AB.match(line) and phase in ('got_name',):
            current['carte_presse'] = clean_carte(line)
            phase = 'got_carte'
            i += 1; continue

        # CP + VILLE
        m_cp = CP_VILLE_RE.match(line)
        if m_cp and phase in ('got_carte', 'got_addr', 'got_name'):
            current['code_postal'] = m_cp.group(1)
            current['ville'] = m_cp.group(2).strip()
            phase = 'got_cp'
            i += 1; continue

        # Adresse
        if phase in ('got_carte', 'got_addr', 'got_name'):
            if not current['adresse']:
                current['adresse'] = line
            else:
                current['adresse'] += ', ' + line
            phase = 'got_addr'

        i += 1

    if current and current.get('nom'):
        fiches.append(current)
    return fiches


# ═════════════════════════════════════════════════════════════
# 6. FORMAT C — 2003+
# ═════════════════════════════════════════════════════════════

EMAIL_PREFIX_STRIP = re.compile(r'^(?:El|K[lp]?|SI|DSl|O|#[=*»])(?:bu|pr)\s*', re.I)

def _clean_email_c(text):
    cleaned = EMAIL_PREFIX_STRIP.sub('', text)
    m = EMAIL_RE.search(cleaned)
    return m.group() if m else None

def _parse_name_c(line):
    m = CARTE_RE_C.search(line)
    if not m:
        return None
    carte = m.group()
    name_part = re.sub(r'\s*[~•]$', '', line[:m.start()]).strip()
    name_part = clean_ocr_icons(name_part)
    if not name_part or len(name_part.split()) < 2:
        return None
    first = re.sub(r'[^A-Za-zÉÈÊËÀÂÙÛÜÔÏÎéèêëàâùûüôïî\-\'\(\)]', '', name_part.split()[0])
    if not first or not first[0].isupper():
        return None
    if NAME_BLACKLIST_STARTS.match(name_part):
        return None
    nom, prenom = split_nom_prenom(name_part)
    prenom = clean_prenom(prenom)
    after = line[m.end():].strip()
    specs = [re.sub(r'[^A-Z]', '', t) for t in re.split(r'[\s/]+', after)]
    spec = '/'.join(s for s in specs if s in SPEC_TAGS)
    return nom, prenom, carte, spec

def _parse_data_segments_c(raw_text):
    result = new_fiche()
    text = raw_text
    placeholders = {}
    for idx, m in enumerate(EMAIL_RE.finditer(text)):
        ph = f'__EM{idx}__'
        placeholders[ph] = m.group()
    for ph, em in placeholders.items():
        text = text.replace(em, ph)
    text = fix_spacing(text)
    for ph, em in placeholders.items():
        text = text.replace(ph, em)

    segments = [s.strip() for s in text.split('•') if s.strip()]
    info_segs, tel_bu, tel_pr, fax_l, gsm_l, email_l = [], [], [], [], [], []

    for seg in segments:
        s = seg.strip()
        if not s: continue
        if re.match(r'^[©®]bu', s):
            tel_bu.extend(PHONE_RE.findall(s)); continue
        if re.match(r'^[©®]pr', s):
            tel_pr.extend(PHONE_RE.findall(s)); continue
        if re.match(r'^©gsm', s):
            gsm_l.extend(PHONE_RE.findall(s)); continue
        if re.match(r'^[\^#<■\\]|^•p', s):
            em = _clean_email_c(s)
            if em: email_l.append(em)
            else: fax_l.extend(PHONE_RE.findall(s))
            continue
        if re.match(r'^(?:El|K[lp]?|SI|DSl|O)(?:bu|pr)', s):
            em = _clean_email_c(s)
            if em: email_l.append(em)
            continue
        info_segs.append(s)

    for seg in info_segs:
        em_m = EMAIL_RE.search(seg)
        if em_m:
            email_l.append(em_m.group())
            seg = (seg[:em_m.start()] + seg[em_m.end():]).strip()
        m_cp = CP_VILLE_RE.search(seg)
        if m_cp and not result['code_postal']:
            before = seg[:m_cp.start()].strip().rstrip(',')
            if before and not result['adresse']:
                result['adresse'] = before
            result['code_postal'] = m_cp.group(1)
            result['ville'] = m_cp.group(2).strip().rstrip('.')
            continue
        is_statut = False
        for st in STATUTS:
            if seg.startswith(st):
                result['journal_ou_statut'] = st
                tel_bu.extend(PHONE_RE.findall(seg[len(st):]))
                is_statut = True; break
        if is_statut: continue
        has_num = bool(re.search(r'\d', seg))
        if not result['adresse'] and not result['code_postal'] and has_num:
            result['adresse'] = seg.rstrip(',').strip()
        elif not result['journal_ou_statut']:
            result['journal_ou_statut'] = seg.strip()
        elif not result['adresse'] and has_num:
            result['adresse'] = seg.rstrip(',').strip()

    if tel_bu: result['tel_bureau'] = tel_bu[0]
    if tel_pr: result['tel_prive'] = tel_pr[0]
    if fax_l:  result['fax'] = fax_l[0]
    if gsm_l:  result['gsm'] = gsm_l[0]
    if email_l: result['email'] = email_l[0]
    return result

_SKIP_C = [re.compile(p) for p in [
    r'^Journalistes$', r'^Listes?$', r'^R[eé]dacteurs$',
    r'^1/1$', r'^CD$', r'^[+\-=ro1c]$', r'^ro\s', r'^[=\-]{2,}',
]]

def _should_skip_c(line):
    if re.match(r'^[FN]?-?\d{1,3}$', line): return True
    if len(line) <= 3 and not any(c.isalpha() for c in line): return True
    return any(p.match(line) for p in _SKIP_C)

def parse_format_c(lines, debug=False):
    groups, cur_name, cur_data = [], None, []
    for line in lines:
        if _should_skip_c(line): continue
        if VOIR_RE.match(line): continue
        parsed = _parse_name_c(line)
        if parsed:
            if cur_name is not None:
                groups.append((cur_name, cur_data))
            cur_name, cur_data = parsed, []
            if debug:
                print(f"  >>> NOM: {parsed[0]} {parsed[1]} [{parsed[2]}] spec={parsed[3]}")
        else:
            cur_data.append(line)
            if debug: print(f"      data: {line}")
    if cur_name is not None:
        groups.append((cur_name, cur_data))

    fiches = []
    for (nom, prenom, carte, spec), data_lines in groups:
        fields = _parse_data_segments_c(' '.join(data_lines))
        fields['nom'] = nom
        fields['prenom'] = prenom
        fields['carte_presse'] = carte
        fields['specialisation'] = spec
        fiches.append(fields)
    return fiches


# ═════════════════════════════════════════════════════════════
# 7. AUTO-DÉTECTION DU FORMAT (vote multi-pages)
# ═════════════════════════════════════════════════════════════

def detect_format(pdf, page_set=None):
    total = len(pdf.pages)
    if page_set:
        test_pages = sorted(page_set)[:min(7, len(page_set))]
    else:
        step = max(1, total // 5)
        test_pages = list(range(min(5, total - 1), total, step))[:5]
        if not test_pages:
            test_pages = [min(10, total - 1)]

    votes = Counter()
    for p in test_pages:
        idx = (p - 1) if page_set else p
        if idx < 0 or idx >= total: continue
        page = pdf.pages[idx]
        text = page.extract_text(x_tolerance=3, y_tolerance=3) or ''
        if '•' in text and text.count('•') > 3:
            votes['c'] += 1
        elif re.search(r'T[eé]l\.\s*(?:bu|pr)\.', text):
            votes['b'] += 1
        else:
            votes['a'] += 1

    if not votes: return 'a'
    winner = votes.most_common(1)[0][0]
    print(f"  Format votes: {dict(votes)} → {winner.upper()}")
    return winner


# ═════════════════════════════════════════════════════════════
# 8. PIPELINE
# ═════════════════════════════════════════════════════════════

PARSERS = {
    'a': (extract_two_columns, parse_format_a),
    'b': (extract_two_columns, parse_format_b),
    'c': (extract_single_column, parse_format_c),
}

def process_pdf(pdf_path, fmt=None, page_ranges=None, label=None, debug=False):
    pdf_path = Path(pdf_path)
    label = label or pdf_path.stem
    all_fiches = []

    with pdfplumber.open(pdf_path) as pdf:
        total = len(pdf.pages)
        page_set = None

        if page_ranges:
            page_set = set()
            for start, end in page_ranges:
                for p in range(max(1, start), min(total, end) + 1):
                    page_set.add(p)
            pages_iter = [(p, pdf.pages[p - 1]) for p in sorted(page_set)]
            desc = ','.join(f'{s}-{e}' for s, e in page_ranges)
            print(f"[{pdf_path.name}] Pages {desc} ({len(pages_iter)} pages)")
        else:
            pages_iter = [(p + 1, page) for p, page in enumerate(pdf.pages)]
            print(f"[{pdf_path.name}] Toutes les pages ({total})")

        if fmt is None:
            fmt = detect_format(pdf, page_set=page_set)
            print(f"  Format détecté : {fmt.upper()}")

        extract_fn, parse_fn = PARSERS[fmt]

        for page_num, page in pages_iter:
            try:
                lines = extract_fn(page)
                fiches = parse_fn(lines, debug=debug)
            except Exception as e:
                print(f"  p.{page_num:>4d} ⚠ ERREUR: {e}")
                continue
            if fiches:
                for f in fiches:
                    f['annuaire'] = label
                    f['page'] = page_num
                all_fiches.extend(fiches)
                if not debug:
                    print(f"  p.{page_num:>4d} → {len(fiches):>3d} fiches")

    n = len(all_fiches)
    if n:
        no_carte = sum(1 for f in all_fiches if not f.get('carte_presse'))
        no_cp    = sum(1 for f in all_fiches if not f.get('code_postal'))
        no_jour  = sum(1 for f in all_fiches if not f.get('journal_ou_statut'))
        no_nom   = sum(1 for f in all_fiches if not f.get('nom'))
        print(f"  TOTAL : {n} fiches")
        if no_nom or no_carte or no_cp or no_jour:
            print(f"  ⚠ Qualité : {no_nom} sans nom, {no_carte} sans carte, "
                  f"{no_cp} sans CP, {no_jour} sans journal")
    else:
        print(f"  TOTAL : 0 fiches")
    print()
    return all_fiches


# ═════════════════════════════════════════════════════════════
# 9. EXPORT EXCEL
# ═════════════════════════════════════════════════════════════

def fiches_to_excel(fiches, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Journalistes"

    headers = [
        "Nom", "Prénom", "N° Carte", "Spécialisation",
        "Adresse", "Code Postal", "Ville", "Journal / Statut",
        "Tél. Bureau", "Tél. Privé", "Fax", "GSM", "Email",
        "Annuaire", "Page",
    ]
    keys = [
        'nom', 'prenom', 'carte_presse', 'specialisation',
        'adresse', 'code_postal', 'ville', 'journal_ou_statut',
        'tel_bureau', 'tel_prive', 'fax', 'gsm', 'email',
        'annuaire', 'page',
    ]
    widths = [22, 18, 10, 12, 35, 10, 25, 30, 18, 18, 18, 18, 30, 12, 6]

    hf = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    hfi = PatternFill("solid", fgColor="2F5496")
    ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cf = Font(name="Arial", size=10)
    ca = Alignment(vertical="top", wrap_text=True)
    bd = Border(*(Side('thin'),) * 4)
    zebra = PatternFill("solid", fgColor="D6E4F0")

    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font, cell.fill, cell.alignment, cell.border = hf, hfi, ha, bd

    for ri, f in enumerate(fiches, 2):
        for ci, k in enumerate(keys, 1):
            cell = ws.cell(ri, ci, f.get(k, ''))
            cell.font, cell.alignment, cell.border = cf, ca, bd
            if ri % 2 == 1:
                cell.fill = zebra

    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(fiches) + 1}"
    ws.freeze_panes = "A2"

    has_spec = any(f.get('specialisation', '') for f in fiches)
    if not has_spec:
        ws.column_dimensions['D'].hidden = True

    wb.save(output_path)
    print(f"Excel exporté : {output_path} ({len(fiches)} fiches)")


# ═════════════════════════════════════════════════════════════
# 10. MAIN
# ═════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="Parser unifié — Annuaires de la Presse Belge (1954-2024) — v2"
    )
    parser.add_argument('pdfs', nargs='+', help="Fichier(s) PDF")
    parser.add_argument('--format', '-f', choices=['a', 'b', 'c'], default=None)
    parser.add_argument('--pages', type=str, default=None,
                        help="Plage(s) de pages, ex: 155-390,560-800")
    parser.add_argument('--label', type=str, default=None)
    parser.add_argument('--output', '-o', type=str, default='annuaire_fiches.xlsx')
    parser.add_argument('--dico', type=str, default=None,
                        help="Dictionnaire médias JSON (nettoyage auto)")
    parser.add_argument('--debug', action='store_true')
    parser.add_argument('--dry-run', action='store_true',
                        help="Analyser sans écrire d'Excel")
    args = parser.parse_args()

    logging.basicConfig(
        level=logging.DEBUG if args.debug else logging.INFO,
        format='%(levelname)s: %(message)s'
    )

    if args.dico:
        if not Path(args.dico).exists():
            print(f"ERREUR: Dictionnaire introuvable : {args.dico}")
            sys.exit(1)

    page_ranges = None
    if args.pages:
        page_ranges = []
        for part in args.pages.split(','):
            se = part.strip().split('-')
            if len(se) == 2:
                page_ranges.append((int(se[0]), int(se[1])))
            else:
                print(f"Format invalide : {part}")
                sys.exit(1)

    all_fiches = []
    for pdf_path in args.pdfs:
        fiches = process_pdf(
            pdf_path, fmt=args.format, page_ranges=page_ranges,
            label=args.label, debug=args.debug,
        )
        all_fiches.extend(fiches)

    if all_fiches and args.dico:
        from nettoyeur import nettoyer_fiches
        all_fiches = nettoyer_fiches(all_fiches, args.dico)

    if args.dry_run:
        print(f"[dry-run] {len(all_fiches)} fiches extraites, pas d'export.")
    elif all_fiches:
        fiches_to_excel(all_fiches, args.output)
    else:
        print("Aucune fiche trouvée !")


if __name__ == "__main__":
    main()
